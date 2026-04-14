import streamlit as st
from openpyxl import load_workbook
from pathlib import Path
import locale


def load_css(file_name):
    """Carrega o arquivo CSS externo."""
    with open(file_name) as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)


def show():
    st.subheader("Conferindo Balanço")

    # Configurar o locale para usar formato monetário brasileiro
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

    # Carregar o CSS externo
    load_css("src/app/css/signal.css")

    # Caminho correto para o arquivo balanco.xlsx
    current_dir = Path(__file__).resolve().parent
    file_path = current_dir.parents[1] / 'balance' / 'balanco.xlsx'

    try:
        # Carregar o arquivo Excel
        workbook = load_workbook(file_path)
        saldo_classificacao_01 = 0
        saldo_classificacao_02 = 0

        # Obter a primeira planilha
        sheet = workbook.active

        # Lendo a primeira linha (células A1 até L1)
        primeira_linha = [sheet.cell(
            row=1, column=col).value for col in range(1, 13)]

        # Filtrar valores vazios (None) da primeira linha antes de exibir no título
        primeira_linha = [str(item)
                          for item in primeira_linha if item is not None]

        # Exibindo a primeira linha no título do Streamlit sem as células vazias
        st.title(f"Balanço: {', '.join(primeira_linha)}")

        # Lendo a segunda linha (células A2 até K2 e L2 até N2)
        segunda_linha_1 = [sheet.cell(
            row=2, column=col).value for col in range(1, 12)]  # A2:K2
        segunda_linha_2 = [sheet.cell(
            row=2, column=col).value for col in range(12, 15)]  # L2:N2

        # Filtrar valores vazios (None) das duas partes da segunda linha antes de exibir no subheader
        segunda_linha = [str(item) for item in segunda_linha_1 +
                         segunda_linha_2 if item is not None]

        # Exibindo a segunda linha no subheader do Streamlit sem as células vazias
        st.subheader(f"Detalhes: {', '.join(segunda_linha)}")

        # Determinar as colunas com base no cabeçalho
        header = [cell.value for cell in sheet[3]]
        classificacao_col = header.index("Classificação") + 1
        saldo_atual_col = header.index("Saldo atual") + 1

        # Iterar pelas linhas e buscar os saldos das classificações "01" e "02"
        for row in sheet.iter_rows(min_row=4, values_only=True):
            classificacao = row[classificacao_col - 1]
            saldo_atual = row[saldo_atual_col - 1]

            # Verificar se o saldo_atual é um número e adicionar corretamente
            try:
                saldo_atual_valor = locale.atof(
                    str(saldo_atual)) if saldo_atual is not None else 0
            except ValueError:
                saldo_atual_valor = 0

            if classificacao == "01":
                saldo_classificacao_01 += saldo_atual_valor
            elif classificacao == "02":
                saldo_classificacao_02 += saldo_atual_valor

        # Formatar os resultados como moeda brasileira com "R$"
        saldo_classificacao_01_formatado = locale.currency(
            saldo_classificacao_01, grouping=True)
        saldo_classificacao_02_formatado = locale.currency(
            saldo_classificacao_02, grouping=True)

        # Exibir os resultados
        st.write(f"Saldo atual para Classificação '01': {
                 saldo_classificacao_01_formatado}")
        st.write(f"Saldo atual para Classificação '02': {
                 saldo_classificacao_02_formatado}")

        # Comparar os saldos
        if saldo_classificacao_01 == saldo_classificacao_02:
            cor = "green"
            mensagem = "Ok, balanço aprovado !"
        else:
            cor = "red"
            mensagem = "Reprovado ! O balanço possui erros ! A operadora estará sendo notificada !"

        # Exibir a bolinha e a mensagem
        st.markdown(f'<div class="status-indicator {cor}"><div class="circle"></div><span>{
                    mensagem}</span></div>', unsafe_allow_html=True)

    except FileNotFoundError as e:
        st.write(f"Erro ao abrir o arquivo: {e}")
    except Exception as e:
        st.write(f"Ocorreu um erro inesperado: {e}")


if __name__ == "__main__":
    show()
