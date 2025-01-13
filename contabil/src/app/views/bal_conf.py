import streamlit as st
from openpyxl import load_workbook
from pathlib import Path
import locale


def show():
    st.subheader("Conferindo Balanço")

    # Configurar o locale para usar formato monetário brasileiro
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

    # Caminho correto para o arquivo balanco.xlsx
    current_dir = Path(__file__).resolve().parent
    file_path = current_dir.parents[1] / 'balance' / 'balanco.xlsx'

    try:
        # Carregar o arquivo Excel
        workbook = load_workbook(file_path)
        saldo_classificacao_01 = 0
        saldo_classificacao_02 = 0

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Determinar as colunas com base no cabeçalho
            header = [cell.value for cell in sheet[3]]
            classificacao_col = header.index("Classificação") + 1
            saldo_atual_col = header.index("Saldo atual") + 1

            # Iterar pelas linhas e buscar os saldos das classificações "01" e "02"
            for row in sheet.iter_rows(min_row=4, values_only=True):
                classificacao = row[classificacao_col - 1]
                saldo_atual = row[saldo_atual_col - 1]

                if classificacao == "01":
                    saldo_classificacao_01 += locale.atof(str(saldo_atual))
                elif classificacao == "02":
                    saldo_classificacao_02 += locale.atof(str(saldo_atual))

        # Formatar os resultados como moeda brasileira com "R$"
        saldo_classificacao_01_formatado = locale.currency(
            saldo_classificacao_01, grouping=True)
        saldo_classificacao_02_formatado = locale.currency(
            saldo_classificacao_02, grouping=True)

        # Exibir os resultados
        st.write(f"Saldo atual para Classificação '01': {
                 saldo_classificacao_01_formatado}")
        st.write(f"Saldo atual para Classificação '02': {
                 saldo_classificacao_02_formatado}")

        # Comparar os saldos
        if saldo_classificacao_01 == saldo_classificacao_02:
            st.write("Ok balanço aprovado")
        else:
            st.write(
                "Reprovado! O balanço possui erros! Enviando notificação para a operadora.")

    except FileNotFoundError as e:
        st.write(f"Erro ao abrir o arquivo: {e}")
    except Exception as e:
        st.write(f"Ocorreu um erro inesperado: {e}")


if __name__ == "__main__":
    show()
