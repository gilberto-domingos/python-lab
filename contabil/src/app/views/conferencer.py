import streamlit as st
from openpyxl import load_workbook
import locale
import subprocess
from dotenv import load_dotenv


load_dotenv()  # Carrega variáveis de ambiente


def load_css(file_name):
    """Carrega o arquivo CSS externo."""
    with open(file_name) as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)


def safe_convert_to_float(value):
    """Converte um valor para float de forma segura, removendo formatações indesejadas."""
    try:
        # Remove espaços e substitui separadores de milhares
        clean_value = str(value).strip().replace(".", "").replace(",", ".")
        return locale.atof(clean_value)
    except (ValueError, TypeError):
        return 0.0


def show():
    st.subheader("Conferindo Balanço")

    # Configurar o locale para usar formato monetário brasileiro
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

    # Carregar o CSS externo
    load_css("src/app/css/signal.css")

    # Criar o uploader de arquivo
    uploaded_file = st.file_uploader("Escolha o arquivo Excel", type=["xlsx"])

    if uploaded_file is not None:
        try:
            # Carregar o arquivo Excel a partir do upload
            workbook = load_workbook(uploaded_file)

            saldo_classificacao_01 = 0
            saldo_classificacao_02 = 0

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Determinar as colunas com base no cabeçalho
                header = [cell.value for cell in sheet[3]]
                classificacao_col = header.index("Classificação") + 1
                saldo_atual_col = header.index("Saldo atual") + 1

                # Iterar pelas linhas e buscar os saldos das classificações "01" e "02"
                for row in sheet.iter_rows(min_row=4, values_only=True):
                    classificacao = str(row[classificacao_col - 1]).strip()
                    saldo_atual = row[saldo_atual_col - 1]

                    if classificacao == "01":
                        saldo_classificacao_01 += safe_convert_to_float(
                            saldo_atual)
                    elif classificacao == "02":
                        saldo_classificacao_02 += safe_convert_to_float(
                            saldo_atual)

            # Formatar os resultados como moeda brasileira com "R$"
            saldo_classificacao_01_formatado = locale.currency(
                saldo_classificacao_01, grouping=True)
            saldo_classificacao_02_formatado = locale.currency(
                saldo_classificacao_02, grouping=True)

            # Exibir os resultados
            st.write(
                f"Saldo atual para Classificação '01': {saldo_classificacao_01_formatado}")
            st.write(
                f"Saldo atual para Classificação '02': {saldo_classificacao_02_formatado}")

            # Comparar os saldos
            if saldo_classificacao_01 == saldo_classificacao_02:
                cor = "green"
                mensagem = "Ok ! balanço aprovado !"
            else:
                cor = "red"
                mensagem = "Reprovado ! O balanço possui erros ! A operadora está sendo notificada !"

                # Exibir a bolinha e a mensagem
                st.markdown(
                    f'<div class="status-error-ok {cor}">'
                    f'<div class="circle"></div><span>{mensagem}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                # Quando o balanço for reprovado, executar o publisher.py e o consumer.py
                try:
                    # Executando o publisher.py
                    subprocess.run(
                        ["python3", "rabbitMq/publisher.py"], check=True)
                    # Executando o consumer.py
                    subprocess.run(
                        ["python3", "rabbitMq/consumer.py"], check=True)
                except subprocess.CalledProcessError as e:
                    st.write(
                        f"Ocorreu um erro ao executar os scripts RabbitMQ: {e}")

        except Exception as e:
            st.write(f"Ocorreu um erro ao processar o arquivo: {e}")


if __name__ == "__main__":
    show()
